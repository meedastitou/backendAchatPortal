"""
════════════════════════════════════════════════════════════
ROUTER - Demandes de Cotation (RFQ)
════════════════════════════════════════════════════════════
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, date
from io import BytesIO
import requests
from app.auth.dependencies import get_current_user, get_user_famille_filter
from app.database import execute_query, execute_insert, execute_x3_query
from app.schemas.rfq import (
    RFQResponse,
    RFQDetailResponse,
    RFQListResponse,
    LigneCotationResponse,
    StatutRFQ,
    # Création manuelle
    FournisseurSelectionResponse,
    FournisseurSearchResponse,
    DADisponibleResponse,
    DAListResponse,
    ArticleDAResponse,
    ArticlesDAListResponse,
    CreerRFQManuelRequest,
    CreerRFQManuelResponse,
    RFQCreatedResponse
)
import uuid as uuid_lib
from app.config import settings

router = APIRouter(prefix="/rfq", tags=["Demandes de Cotation"])


# ──────────────────────────────────────────────────────────
# Liste des RFQ
# ──────────────────────────────────────────────────────────

@router.get("", response_model=RFQListResponse)
async def list_rfq(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    statut: Optional[StatutRFQ] = None,
    code_fournisseur: Optional[str] = None,
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    search: Optional[str] = None,
    code_article: Optional[str] = None,
    numero_da: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les demandes de cotation avec filtres"""

    conditions = ["1=1"]
    params = []
    join_lignes = False
    join_articles = False

    # Filtrage par famille pour les acheteurs
    familles_filter = get_user_famille_filter(current_user)
    if familles_filter is not None:
        if len(familles_filter) == 0:
            # Acheteur sans famille assignée = ne voit rien
            return RFQListResponse(rfqs=[], total=0, page=page, limit=limit)
        join_lignes = True
        join_articles = True
        placeholders = ", ".join(["%s"] * len(familles_filter))
        conditions.append(f"ar.code_famille IN ({placeholders})")
        params.extend(familles_filter)

    if statut:
        conditions.append("dc.statut = %s")
        params.append(statut.value)

    if code_fournisseur:
        conditions.append("dc.code_fournisseur = %s")
        params.append(code_fournisseur)

    if date_debut:
        conditions.append("dc.date_envoi >= %s")
        params.append(date_debut)

    if date_fin:
        conditions.append("dc.date_envoi <= %s")
        params.append(date_fin)

    if search:
        conditions.append("(dc.numero_rfq LIKE %s OR f.nom_fournisseur LIKE %s)")
        search_pattern = f"%{search}%"
        params.extend([search_pattern, search_pattern])

    if code_article:
        join_lignes = True
        conditions.append("lc.code_article LIKE %s")
        params.append(f"%{code_article}%")

    if numero_da:
        join_lignes = True
        conditions.append("lc.numero_da LIKE %s")
        params.append(f"%{numero_da}%")

    where_clause = " AND ".join(conditions)
    lignes_join = "JOIN lignes_cotation lc ON dc.uuid = lc.rfq_uuid" if join_lignes else ""
    articles_join = "JOIN articles_ref ar ON lc.code_article = ar.code_article" if join_articles else ""

    # Count
    count_query = f"""
        SELECT COUNT(DISTINCT dc.id) as total
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        {lignes_join}
        {articles_join}
        WHERE {where_clause}
    """
    total = execute_query(count_query, tuple(params), fetch_one=True)["total"]

    # Get RFQs
    offset = (page - 1) * limit
    query = f"""
        SELECT DISTINCT
            dc.*,
            f.nom_fournisseur,
            f.email as email_fournisseur
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        {lignes_join}
        {articles_join}
        WHERE {where_clause}
        ORDER BY dc.date_envoi DESC
        LIMIT %s OFFSET %s
    """
    params.extend([limit, offset])
    rfqs = execute_query(query, tuple(params))

    # Ajouter les lignes pour chaque RFQ
    rfq_responses = []
    for rfq in rfqs:
        lignes = execute_query(
            "SELECT * FROM lignes_cotation WHERE rfq_uuid = %s",
            (rfq["uuid"],)
        )
        rfq_responses.append(RFQResponse(
            **rfq,
            lignes=[LigneCotationResponse(**l) for l in lignes]
        ))

    return RFQListResponse(
        rfqs=rfq_responses,
        total=total,
        page=page,
        limit=limit
    )


# ──────────────────────────────────────────────────────────
# Détail d'une RFQ
# ──────────────────────────────────────────────────────────

@router.get("/{rfq_id}", response_model=RFQDetailResponse)
async def get_rfq(
    rfq_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir les détails d'une RFQ"""

    query = """
        SELECT
            dc.*,
            f.nom_fournisseur,
            f.email as email_fournisseur,
            DATEDIFF(NOW(), dc.date_envoi) as jours_depuis_envoi,
            TIMESTAMPDIFF(HOUR, dc.date_envoi, dc.date_reponse) as delai_reponse_heures
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        WHERE dc.id = %s
    """
    rfq = execute_query(query, (rfq_id,), fetch_one=True)

    if not rfq:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="RFQ non trouvée"
        )

    # Vérifier accès par famille pour les acheteurs
    familles_filter = get_user_famille_filter(current_user)
    if familles_filter is not None:
        # Vérifier si au moins un article de cette RFQ appartient aux familles de l'utilisateur
        if len(familles_filter) == 0:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Accès non autorisé à cette RFQ"
            )
        placeholders = ", ".join(["%s"] * len(familles_filter))
        access_check = execute_query(
            f"""
            SELECT 1 FROM lignes_cotation lc
            JOIN articles_ref ar ON lc.code_article = ar.code_article
            WHERE lc.rfq_uuid = %s AND ar.code_famille IN ({placeholders})
            LIMIT 1
            """,
            (rfq["uuid"], *familles_filter),
            fetch_one=True
        )
        if not access_check:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Accès non autorisé à cette RFQ"
            )

    # Récupérer les lignes
    lignes = execute_query(
        "SELECT * FROM lignes_cotation WHERE rfq_uuid = %s",
        (rfq["uuid"],)
    )

    return RFQDetailResponse(
        **rfq,
        lignes=[LigneCotationResponse(**l) for l in lignes],
        nb_articles=len(lignes)
    )


# ──────────────────────────────────────────────────────────
# RFQ par UUID
# ──────────────────────────────────────────────────────────

@router.get("/uuid/{uuid}", response_model=RFQDetailResponse)
async def get_rfq_by_uuid(
    uuid: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir une RFQ par son UUID"""

    query = """
        SELECT
            dc.*,
            f.nom_fournisseur,
            f.email as email_fournisseur,
            DATEDIFF(NOW(), dc.date_envoi) as jours_depuis_envoi,
            TIMESTAMPDIFF(HOUR, dc.date_envoi, dc.date_reponse) as delai_reponse_heures
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        WHERE dc.uuid = %s
    """
    rfq = execute_query(query, (uuid,), fetch_one=True)

    if not rfq:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="RFQ non trouvée"
        )

    lignes = execute_query(
        "SELECT * FROM lignes_cotation WHERE rfq_uuid = %s",
        (uuid,)
    )

    return RFQDetailResponse(
        **rfq,
        lignes=[LigneCotationResponse(**l) for l in lignes],
        nb_articles=len(lignes)
    )


# ──────────────────────────────────────────────────────────
# Statistiques par statut
# ──────────────────────────────────────────────────────────

@router.get("/stats/by-status")
async def get_rfq_stats_by_status(current_user: dict = Depends(get_current_user)):
    """Statistiques des RFQ par statut"""

    query = """
        SELECT statut, COUNT(*) as count
        FROM demandes_cotation
        GROUP BY statut
        ORDER BY count DESC
    """
    results = execute_query(query)

    return {
        "stats": results,
        "total": sum(r["count"] for r in results)
    }


# ──────────────────────────────────────────────────────────
# RFQ en attente de réponse
# ──────────────────────────────────────────────────────────

@router.get("/pending/list")
async def get_pending_rfq(
    days_old: int = Query(7, ge=1),
    current_user: dict = Depends(get_current_user)
):
    """Lister les RFQ en attente depuis plus de X jours"""

    query = """
        SELECT
            dc.*,
            f.nom_fournisseur,
            f.email as email_fournisseur,
            DATEDIFF(NOW(), dc.date_envoi) as jours_depuis_envoi
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        WHERE dc.statut IN ('envoye', 'relance_1', 'relance_2', 'relance_3')
          AND dc.date_reponse IS NULL
          AND DATEDIFF(NOW(), dc.date_envoi) >= %s
        ORDER BY dc.date_envoi ASC
    """
    rfqs = execute_query(query, (days_old,))

    return {
        "rfqs": rfqs,
        "total": len(rfqs),
        "days_threshold": days_old
    }


# ──────────────────────────────────────────────────────────
# Export Excel - RFQ sans réponse
# ──────────────────────────────────────────────────────────

@router.get("/export/sans-reponse")
async def export_rfq_sans_reponse(
    date_debut: Optional[date] = None,
    date_fin: Optional[date] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Exporter les RFQ sans réponse en Excel.
    Statuts inclus: envoye, relance_1, relance_2, relance_3
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Module openpyxl non installé. Exécutez: pip install openpyxl"
        )
    print("test===================")
    # Construire la requête
    conditions = ["dc.statut IN ('envoye', 'relance_1', 'relance_2', 'relance_3')"]
    params = []

    if date_debut:
        conditions.append("DATE(dc.date_envoi) >= %s")
        params.append(date_debut)

    if date_fin:
        conditions.append("DATE(dc.date_envoi) <= %s")
        params.append(date_fin)

    where_clause = " AND ".join(conditions)

    query = f"""
        SELECT
            dc.numero_rfq,
            dc.code_fournisseur,
            f.nom_fournisseur,
            f.email as email_fournisseur,
            dc.date_envoi,
            dc.nb_relances,
            dc.date_derniere_relance,
            dc.uuid,
            DATEDIFF(NOW(), dc.date_envoi) as jours_depuis_envoi
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        WHERE {where_clause}
        ORDER BY dc.date_envoi DESC
    """

    rfqs = execute_query(query, tuple(params) if params else None)

    # Récupérer les articles pour chaque RFQ
    for rfq in rfqs:
        lignes = execute_query(
            "SELECT code_article, numero_da FROM lignes_cotation WHERE rfq_uuid = %s",
            (rfq["uuid"],)
        )
        rfq["articles"] = ", ".join([l["code_article"] for l in lignes])
        rfq["das"] = ", ".join(list(set([l["numero_da"] for l in lignes])))

    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ Sans Réponse"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D5A87", end_color="2D5A87", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = [
        "N° RFQ",
        "Code Fournisseur",
        "Nom Fournisseur",
        "Email",
        "Date Envoi",
        "Nb Relances",
        "Dernière Relance",
        "Jours Depuis Envoi",
        "Articles",
        "N° DA"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Données
    for row_num, rfq in enumerate(rfqs, 2):
        data = [
            rfq["numero_rfq"],
            rfq["code_fournisseur"],
            rfq["nom_fournisseur"],
            rfq["email_fournisseur"],
            rfq["date_envoi"].strftime("%d/%m/%Y %H:%M") if rfq["date_envoi"] else "",
            rfq["nb_relances"],
            rfq["date_derniere_relance"].strftime("%d/%m/%Y") if rfq["date_derniere_relance"] else "",
            rfq["jours_depuis_envoi"],
            rfq["articles"],
            rfq["das"]
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col == 8 and value and value > 7:  # Jours depuis envoi > 7
                cell.font = Font(color="DC2626", bold=True)

    # Ajuster la largeur des colonnes
    column_widths = [15, 18, 30, 35, 18, 12, 18, 18, 40, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Figer la première ligne
    ws.freeze_panes = "A2"

    # Générer le fichier
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nom du fichier
    date_str_debut = date_debut.strftime("%Y-%m-%d") if date_debut else "debut"
    date_str_fin = date_fin.strftime("%Y-%m-%d") if date_fin else "fin"
    filename = f"RFQ_sans_reponse_{date_str_debut}_{date_str_fin}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ──────────────────────────────────────────────────────────
# Export Excel - RFQ avec filtres
# ──────────────────────────────────────────────────────────

@router.get("/export/filtered")
async def export_rfq_filtered(
    statut: Optional[StatutRFQ] = None,
    code_fournisseur: Optional[str] = None,
    date_debut: Optional[date] = None,
    date_fin: Optional[date] = None,
    search: Optional[str] = None,
    code_article: Optional[str] = None,
    numero_da: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Exporter les RFQ en Excel selon les filtres appliqués.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Module openpyxl non installé. Exécutez: pip install openpyxl"
        )

    # Construire la requête avec les mêmes filtres que list_rfq
    conditions = ["1=1"]
    params = []
    join_lignes = False

    if statut:
        conditions.append("dc.statut = %s")
        params.append(statut.value)

    if code_fournisseur:
        conditions.append("dc.code_fournisseur = %s")
        params.append(code_fournisseur)

    if date_debut:
        conditions.append("DATE(dc.date_envoi) >= %s")
        params.append(date_debut)

    if date_fin:
        conditions.append("DATE(dc.date_envoi) <= %s")
        params.append(date_fin)

    if search:
        conditions.append("(dc.numero_rfq LIKE %s OR f.nom_fournisseur LIKE %s)")
        search_pattern = f"%{search}%"
        params.extend([search_pattern, search_pattern])

    if code_article:
        join_lignes = True
        conditions.append("lc.code_article LIKE %s")
        params.append(f"%{code_article}%")

    if numero_da:
        join_lignes = True
        conditions.append("lc.numero_da LIKE %s")
        params.append(f"%{numero_da}%")

    where_clause = " AND ".join(conditions)
    lignes_join = "JOIN lignes_cotation lc ON dc.uuid = lc.rfq_uuid" if join_lignes else ""

    query = f"""
        SELECT DISTINCT
            dc.numero_rfq,
            dc.code_fournisseur,
            f.nom_fournisseur,
            f.email as email_fournisseur,
            dc.date_envoi,
            dc.statut,
            dc.nb_relances,
            dc.date_derniere_relance,
            dc.date_reponse,
            dc.uuid,
            DATEDIFF(NOW(), dc.date_envoi) as jours_depuis_envoi
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        {lignes_join}
        WHERE {where_clause}
        ORDER BY dc.date_envoi DESC
    """

    rfqs = execute_query(query, tuple(params) if params else None)

    # Récupérer les articles pour chaque RFQ
    for rfq in rfqs:
        lignes = execute_query(
            "SELECT code_article, numero_da FROM lignes_cotation WHERE rfq_uuid = %s",
            (rfq["uuid"],)
        )
        rfq["articles"] = ", ".join([l["code_article"] for l in lignes])
        rfq["das"] = ", ".join(list(set([l["numero_da"] for l in lignes])))

    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Export RFQ"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D5A87", end_color="2D5A87", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = [
        "N° RFQ",
        "Code Fournisseur",
        "Nom Fournisseur",
        "Email",
        "Date Envoi",
        "Statut",
        "Nb Relances",
        "Dernière Relance",
        "Date Réponse",
        "Jours Depuis Envoi",
        "Articles",
        "N° DA"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Mapping des statuts pour affichage
    statut_map = {
        'envoye': 'Envoyé',
        'vu': 'Vu',
        'repondu': 'Répondu',
        'rejete': 'Rejeté',
        'expire': 'Expiré',
        'relance_1': 'Relance 1',
        'relance_2': 'Relance 2',
        'relance_3': 'Relance 3'
    }

    # Données
    for row_num, rfq in enumerate(rfqs, 2):
        data = [
            rfq["numero_rfq"],
            rfq["code_fournisseur"],
            rfq["nom_fournisseur"],
            rfq["email_fournisseur"],
            rfq["date_envoi"].strftime("%d/%m/%Y %H:%M") if rfq["date_envoi"] else "",
            statut_map.get(rfq["statut"], rfq["statut"]),
            rfq["nb_relances"],
            rfq["date_derniere_relance"].strftime("%d/%m/%Y") if rfq["date_derniere_relance"] else "",
            rfq["date_reponse"].strftime("%d/%m/%Y %H:%M") if rfq["date_reponse"] else "",
            rfq["jours_depuis_envoi"],
            rfq["articles"],
            rfq["das"]
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

            # Colorer selon le statut
            if col == 6:  # Colonne Statut
                if rfq["statut"] == "repondu":
                    cell.font = Font(color="059669", bold=True)
                elif rfq["statut"] == "rejete":
                    cell.font = Font(color="DC2626", bold=True)
                elif rfq["statut"] in ("relance_1", "relance_2", "relance_3"):
                    cell.font = Font(color="D97706", bold=True)

    # Ajuster la largeur des colonnes
    column_widths = [15, 18, 30, 35, 18, 12, 12, 18, 18, 15, 40, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Figer la première ligne
    ws.freeze_panes = "A2"

    # Générer le fichier
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nom du fichier
    from datetime import datetime as dt
    filename = f"Export_RFQ_{dt.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ──────────────────────────────────────────────────────────
# DA Non Soldées (depuis Sage X3)
# ──────────────────────────────────────────────────────────

@router.get("/da-non-solde/list")
async def get_da_non_soldees(
    current_user: dict = Depends(get_current_user)
):
    """
    Récupérer les DA et articles non soldés depuis Sage X3.
    Requête vers la vue Y_DA_NON_SOLDEES.
    """
    query = """
        SELECT
            DNS.DA AS numero_da,
            DNS.Article AS code_article,
            DNS.DésignationArticle AS designation_article,
            DNS.QteDA AS quantite,
            DAD.XMARQ_0 AS marque,
            DNS.STU_0 AS unite,
            DNS.Famille AS famille,
            DAD.LINAPPFLG_0 AS niveau_signature
        FROM x3.BASE1.Y_DA_NON_SOLDEES DNS
        LEFT JOIN x3.BASE1.PREQUIS DA ON DA.PSHNUM_0 = DNS.DA
        INNER JOIN x3.BASE1.PREQUISD DAD
            ON DAD.PSHNUM_0 = DA.PSHNUM_0 AND DAD.ITMREF_0 = DNS.Article
        WHERE DA.CLEFLG_0 = 1
        ORDER BY DNS.DA, DNS.Article
    """

    try:
        results = execute_x3_query(query)
        return {
            "da_non_soldees": results,
            "total": len(results)
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la récupération des DA non soldées: {str(e)}"
        )


@router.get("/da-non-solde/analyse")
async def get_da_non_solde_analyse(
    current_user: dict = Depends(get_current_user)
):
    """
    Analyse des DA non soldées avec les RFQ associées.
    Retourne:
    - DA/articles sans RFQ
    - RFQ sans réponse
    - RFQ avec réponse
    """
    # 1. Récupérer les DA non soldées depuis X3
    x3_query = """
        SELECT
            DNS.DA, DNS.Article, DNS.DésignationArticle, DNS.QteDA,
            DAD.XMARQ_0 AS marque, DNS.STU_0 AS Unite, DNS.Famille AS Famille
        FROM x3.BASE1.Y_DA_NON_SOLDEES DNS
        LEFT JOIN x3.BASE1.PREQUIS PRQ ON PRQ.PSHNUM_0 = DNS.DA
        INNER JOIN x3.BASE1.PREQUISD DAD
            ON DAD.PSHNUM_0 = PRQ.PSHNUM_0 AND DAD.ITMREF_0 = DNS.Article
        WHERE PRQ.CLEFLG_0 = 1
        ORDER BY DNS.DA, DNS.Article
    """

    try:
        da_non_soldees = execute_x3_query(x3_query)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur X3: {str(e)}"
        )

    if not da_non_soldees:
        return {
            "da_non_soldees": [],
            "total_da": 0,
            "da_sans_rfq": [],
            "total_sans_rfq": 0,
            "rfqs_sans_reponse": [],
            "total_sans_reponse": 0,
            "rfqs_avec_reponse": [],
            "total_avec_reponse": 0
        }

    # 2. Identifier les DA/articles qui ont une RFQ (dans lignes_cotation)
    # Récupérer toutes les combinaisons DA/article ayant une RFQ
    da_avec_rfq_query = """
        SELECT DISTINCT numero_da, code_article
        FROM lignes_cotation
    """
    da_avec_rfq_list = execute_query(da_avec_rfq_query)
    da_avec_rfq_set = {(item["numero_da"], item["code_article"]) for item in da_avec_rfq_list}

    # 3. Séparer les DA non soldées en deux catégories
    da_sans_rfq = []
    da_avec_rfq = []

    for da in da_non_soldees:
        key = (da["DA"], da["Article"])
        if key in da_avec_rfq_set:
            da_avec_rfq.append(da)
        else:
            da_sans_rfq.append(da)

    # Si aucune DA n'a de RFQ, retourner
    if not da_avec_rfq:
        return {
            "da_non_soldees": da_non_soldees,
            "total_da": len(da_non_soldees),
            "da_sans_rfq": da_sans_rfq,
            "total_sans_rfq": len(da_sans_rfq),
            "rfqs_sans_reponse": [],
            "total_sans_reponse": 0,
            "rfqs_avec_reponse": [],
            "total_avec_reponse": 0
        }

    # 4. Construire la condition pour matcher les DA/articles avec RFQ
    da_article_conditions = []
    params = []
    for item in da_avec_rfq:
        da_article_conditions.append("(lc.numero_da = %s AND lc.code_article = %s)")
        params.append(item["DA"])
        params.append(item["Article"])

    da_filter = f"({' OR '.join(da_article_conditions)})"

    # 5. Récupérer les RFQ SANS réponse (statut: envoye, relance_1, relance_2, relance_3)
    query_sans_reponse = f"""
        SELECT DISTINCT
            dc.id,
            dc.uuid,
            dc.numero_rfq,
            dc.code_fournisseur,
            f.nom_fournisseur,
            f.email as email_fournisseur,
            dc.date_envoi,
            dc.statut,
            dc.nb_relances,
            dc.date_derniere_relance,
            DATEDIFF(NOW(), dc.date_envoi) as jours_depuis_envoi
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        JOIN lignes_cotation lc ON dc.uuid = lc.rfq_uuid
        WHERE {da_filter}
          AND dc.statut IN ('envoye', 'relance_1', 'relance_2', 'relance_3')
        ORDER BY dc.date_envoi DESC
    """
    rfqs_sans_reponse = execute_query(query_sans_reponse, tuple(params))

    # Ajouter les lignes pour chaque RFQ sans réponse
    for rfq in rfqs_sans_reponse:
        lignes = execute_query(
            "SELECT numero_da, code_article, designation_article, quantite_demandee, unite FROM lignes_cotation WHERE rfq_uuid = %s",
            (rfq["uuid"],)
        )
        rfq["lignes"] = lignes
        rfq["nb_articles"] = len(lignes)

    # 6. Récupérer les RFQ AVEC réponse (statut: repondu)
    query_avec_reponse = f"""
        SELECT DISTINCT
            dc.id,
            dc.uuid,
            dc.numero_rfq,
            dc.code_fournisseur,
            f.nom_fournisseur,
            f.email as email_fournisseur,
            dc.date_envoi,
            dc.date_reponse,
            dc.statut,
            TIMESTAMPDIFF(HOUR, dc.date_envoi, dc.date_reponse) as delai_reponse_heures
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        JOIN lignes_cotation lc ON dc.uuid = lc.rfq_uuid
        WHERE {da_filter}
          AND dc.statut = 'repondu'
        ORDER BY dc.date_reponse DESC
    """
    rfqs_avec_reponse = execute_query(query_avec_reponse, tuple(params))

    # Ajouter les lignes pour chaque RFQ avec réponse
    for rfq in rfqs_avec_reponse:
        lignes = execute_query(
            "SELECT numero_da, code_article, designation_article, quantite_demandee, unite FROM lignes_cotation WHERE rfq_uuid = %s",
            (rfq["uuid"],)
        )
        rfq["lignes"] = lignes
        rfq["nb_articles"] = len(lignes)

    return {
        "da_non_soldees": da_non_soldees,
        "total_da": len(da_non_soldees),
        "da_sans_rfq": da_sans_rfq,
        "total_sans_rfq": len(da_sans_rfq),
        "rfqs_sans_reponse": rfqs_sans_reponse,
        "total_sans_reponse": len(rfqs_sans_reponse),
        "rfqs_avec_reponse": rfqs_avec_reponse,
        "total_avec_reponse": len(rfqs_avec_reponse)
    }


@router.get("/da-non-solde/rfqs", response_model=RFQListResponse)
async def list_rfq_da_non_solde(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    statut: Optional[StatutRFQ] = None,
    code_fournisseur: Optional[str] = None,
    date_debut: Optional[datetime] = None,
    date_fin: Optional[datetime] = None,
    search: Optional[str] = None,
    code_article: Optional[str] = None,
    numero_da: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Lister les RFQ qui correspondent aux DA non soldées de X3.
    Filtre les RFQ dont les lignes de cotation matchent les DA/articles non soldés.
    """

    # 1. Récupérer les DA non soldées depuis X3
    x3_query = """
        SELECT
            DNS.DA AS numero_da,
            DNS.Article AS code_article
        FROM x3.BASE1.Y_DA_NON_SOLDEES DNS
        LEFT JOIN x3.BASE1.PREQUIS DA ON DA.PSHNUM_0 = DNS.DA
        INNER JOIN x3.BASE1.PREQUISD DAD
            ON DAD.PSHNUM_0 = DA.PSHNUM_0 AND DAD.ITMREF_0 = DNS.Article
        WHERE DA.CLEFLG_0 = 1
    """

    try:
        da_non_soldees = execute_x3_query(x3_query)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur X3: {str(e)}"
        )

    if not da_non_soldees:
        return RFQListResponse(rfqs=[], total=0, page=page, limit=limit)

    # 2. Construire les conditions pour filtrer les RFQ
    conditions = ["1=1"]
    params = []
    join_lignes = True  # Toujours requis pour ce filtre
    join_articles = False

    # Créer la condition pour matcher les DA/articles non soldés
    da_article_conditions = []
    for item in da_non_soldees:
        da_article_conditions.append("(lc.numero_da = %s AND lc.code_article = %s)")
        params.append(item["numero_da"])
        params.append(item["code_article"])

    if da_article_conditions:
        conditions.append(f"({' OR '.join(da_article_conditions)})")

    # Filtrage par famille pour les acheteurs
    familles_filter = get_user_famille_filter(current_user)
    if familles_filter is not None:
        if len(familles_filter) == 0:
            return RFQListResponse(rfqs=[], total=0, page=page, limit=limit)
        join_articles = True
        placeholders = ", ".join(["%s"] * len(familles_filter))
        conditions.append(f"ar.code_famille IN ({placeholders})")
        params.extend(familles_filter)

    if statut:
        conditions.append("dc.statut = %s")
        params.append(statut.value)

    if code_fournisseur:
        conditions.append("dc.code_fournisseur = %s")
        params.append(code_fournisseur)

    if date_debut:
        conditions.append("dc.date_envoi >= %s")
        params.append(date_debut)

    if date_fin:
        conditions.append("dc.date_envoi <= %s")
        params.append(date_fin)

    if search:
        conditions.append("(dc.numero_rfq LIKE %s OR f.nom_fournisseur LIKE %s)")
        search_pattern = f"%{search}%"
        params.extend([search_pattern, search_pattern])

    if code_article:
        conditions.append("lc.code_article LIKE %s")
        params.append(f"%{code_article}%")

    if numero_da:
        conditions.append("lc.numero_da LIKE %s")
        params.append(f"%{numero_da}%")

    where_clause = " AND ".join(conditions)
    lignes_join = "JOIN lignes_cotation lc ON dc.uuid = lc.rfq_uuid"
    articles_join = "JOIN articles_ref ar ON lc.code_article = ar.code_article" if join_articles else ""

    # Count
    count_query = f"""
        SELECT COUNT(DISTINCT dc.id) as total
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        {lignes_join}
        {articles_join}
        WHERE {where_clause}
    """
    total = execute_query(count_query, tuple(params), fetch_one=True)["total"]

    # Get RFQs
    offset = (page - 1) * limit
    query = f"""
        SELECT DISTINCT
            dc.*,
            f.nom_fournisseur,
            f.email as email_fournisseur
        FROM demandes_cotation dc
        JOIN fournisseurs f ON dc.code_fournisseur = f.code_fournisseur
        {lignes_join}
        {articles_join}
        WHERE {where_clause}
        ORDER BY dc.date_envoi DESC
        LIMIT %s OFFSET %s
    """
    params.extend([limit, offset])
    rfqs = execute_query(query, tuple(params))

    # Ajouter les lignes pour chaque RFQ
    rfq_responses = []
    for rfq in rfqs:
        lignes = execute_query(
            "SELECT * FROM lignes_cotation WHERE rfq_uuid = %s",
            (rfq["uuid"],)
        )
        rfq_responses.append(RFQResponse(
            **rfq,
            lignes=[LigneCotationResponse(**l) for l in lignes]
        ))

    return RFQListResponse(
        rfqs=rfq_responses,
        total=total,
        page=page,
        limit=limit
    )


# ══════════════════════════════════════════════════════════════
# CRÉATION MANUELLE DE RFQ
# ══════════════════════════════════════════════════════════════

# ──────────────────────────────────────────────────────────
# Recherche de fournisseurs
# ──────────────────────────────────────────────────────────

@router.get("/creation/fournisseurs", response_model=FournisseurSearchResponse)
async def search_fournisseurs_for_rfq(
    q: str = Query("", description="Recherche par code ou nom"),
    limit: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """Rechercher des fournisseurs pour la création de RFQ"""

    conditions = ["blacklist = FALSE"]
    params = []

    if q:
        conditions.append("(code_fournisseur LIKE %s OR nom_fournisseur LIKE %s)")
        search_pattern = f"%{q}%"
        params.extend([search_pattern, search_pattern])

    where_clause = " AND ".join(conditions)

    # Count
    count_query = f"SELECT COUNT(*) as total FROM fournisseurs WHERE {where_clause}"
    total = execute_query(count_query, tuple(params) if params else None, fetch_one=True)["total"]

    # Get fournisseurs
    query = f"""
        SELECT code_fournisseur, nom_fournisseur, email, blacklist
        FROM fournisseurs
        WHERE {where_clause}
        ORDER BY nom_fournisseur ASC
        LIMIT %s
    """
    params.append(limit)
    fournisseurs = execute_query(query, tuple(params))

    return FournisseurSearchResponse(
        fournisseurs=[FournisseurSelectionResponse(**f) for f in fournisseurs],
        total=total
    )


# ──────────────────────────────────────────────────────────
# DA disponibles (depuis demandes_achat MySQL)
# ──────────────────────────────────────────────────────────

@router.get("/creation/da-disponibles", response_model=DAListResponse)
async def get_da_disponibles_for_rfq(
    search: str = Query("", description="Recherche par numéro DA"),
    current_user: dict = Depends(get_current_user)
):
    """Lister les DA disponibles depuis demandes_achat (MySQL)"""

    conditions = ["statut IN ('nouveau', 'en_cours')"]
    params = []

    if search:
        conditions.append("numero_da LIKE %s")
        params.append(f"%{search}%")

    where_clause = " AND ".join(conditions)

    query = f"""
        SELECT numero_da, COUNT(*) as nb_articles
        FROM demandes_achat
        WHERE {where_clause}
        GROUP BY numero_da
        ORDER BY numero_da DESC
    """

    das = execute_query(query, tuple(params) if params else None)

    return DAListResponse(
        da_list=[DADisponibleResponse(**da) for da in das],
        total=len(das)
    )


# ──────────────────────────────────────────────────────────
# Articles d'une DA
# ──────────────────────────────────────────────────────────

@router.get("/creation/da/{numero_da}/articles", response_model=ArticlesDAListResponse)
async def get_articles_da_for_rfq(
    numero_da: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les articles d'une DA depuis demandes_achat (MySQL)"""

    query = """
        SELECT id, numero_da, code_article, designation_article,
               quantite, unite, marque_souhaitee
        FROM demandes_achat
        WHERE numero_da = %s
        ORDER BY code_article
    """

    articles = execute_query(query, (numero_da,))

    if not articles:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Aucun article trouvé pour la DA {numero_da}"
        )

    return ArticlesDAListResponse(
        numero_da=numero_da,
        articles=[ArticleDAResponse(**a) for a in articles],
        total=len(articles)
    )


# ──────────────────────────────────────────────────────────
# Créer des RFQ manuellement
# ──────────────────────────────────────────────────────────

@router.post("/creation/creer", response_model=CreerRFQManuelResponse)
async def creer_rfq_manuel(
    request: CreerRFQManuelRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Créer manuellement des RFQ (une par fournisseur sélectionné).
    Tous les fournisseurs reçoivent les mêmes articles.
    """

    if not request.fournisseurs:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Au moins un fournisseur doit être sélectionné"
        )

    if not request.articles:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Au moins un article doit être sélectionné"
        )

    # Vérifier que les fournisseurs existent et ne sont pas blacklistés
    fournisseurs_codes = request.fournisseurs
    placeholders = ", ".join(["%s"] * len(fournisseurs_codes))
    fournisseurs_db = execute_query(
        f"""
        SELECT code_fournisseur, nom_fournisseur, email, blacklist
        FROM fournisseurs
        WHERE code_fournisseur IN ({placeholders})
        """,
        tuple(fournisseurs_codes)
    )

    fournisseurs_map = {f["code_fournisseur"]: f for f in fournisseurs_db}

    # Vérifier que tous existent
    for code in fournisseurs_codes:
        if code not in fournisseurs_map:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Fournisseur {code} non trouvé"
            )
        if fournisseurs_map[code]["blacklist"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Fournisseur {code} est blacklisté"
            )

    # Générer le préfixe du numéro RFQ
    today = datetime.now()
    prefix = f"RFQ-{today.strftime('%Y%m%d')}"

    # Récupérer le dernier numéro RFQ du jour
    last_rfq = execute_query(
        "SELECT numero_rfq FROM demandes_cotation WHERE numero_rfq LIKE %s ORDER BY numero_rfq DESC LIMIT 1",
        (f"{prefix}%",),
        fetch_one=True
    )

    if last_rfq:
        # Extraire le compteur
        try:
            last_num = int(last_rfq["numero_rfq"].split("-")[-1])
        except (ValueError, IndexError):
            last_num = 0
    else:
        last_num = 0

    rfqs_crees = []
    nb_emails_envoyes = 0

    # Créer une RFQ par fournisseur
    for i, code_fournisseur in enumerate(fournisseurs_codes):
        fournisseur = fournisseurs_map[code_fournisseur]
        rfq_uuid = str(uuid_lib.uuid4())
        numero_rfq = f"{prefix}-{str(last_num + i + 1).zfill(4)}"

        # Insérer la demande de cotation
        execute_insert(
            """
            INSERT INTO demandes_cotation
            (uuid, numero_rfq, code_fournisseur, date_envoi, date_limite_reponse, statut, nb_relances, manuel, created_by)
            VALUES (%s, %s, %s, %s, %s, 'envoye', 0, 1, %s)
            """,
            (
                rfq_uuid,
                numero_rfq,
                code_fournisseur,
                today,
                request.date_limite_reponse,
                current_user["username"]
            )
        )

        # Insérer les lignes de cotation
        for article in request.articles:
            execute_insert(
                """
                INSERT INTO lignes_cotation
                (rfq_uuid, numero_da, code_article, designation_article, quantite_demandee, unite, marque_souhaitee)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    rfq_uuid,
                    article.numero_da,
                    article.code_article,
                    article.designation_article,
                    article.quantite,
                    article.unite,
                    article.marque_souhaitee
                )
            )

        email_envoye = False
        email_error = None
        response = requests.post(
            url=settings.N8N_endpoint_URL_ACH4,
            json={'rfq_uuid':rfq_uuid, 'email_acheteur': current_user["email"]}
            # headers=default_headers,
            # timeout=self.timeout
        )   
        if(response.status_code == 200):
            email_envoye=True

        rfqs_crees.append(RFQCreatedResponse(
            uuid=rfq_uuid,
            numero_rfq=numero_rfq,
            code_fournisseur=code_fournisseur,
            nom_fournisseur=fournisseur["nom_fournisseur"],
            email=fournisseur["email"],
            nb_articles=len(request.articles),
            email_envoye=email_envoye,
            email_error=email_error
        ))

        if email_envoye:
            nb_emails_envoyes += 1

    return CreerRFQManuelResponse(
        success=True,
        message=f"{len(rfqs_crees)} RFQ créées avec succès",
        rfqs_crees=rfqs_crees,
        nb_rfqs=len(rfqs_crees),
        nb_emails_envoyes=nb_emails_envoyes
    )
